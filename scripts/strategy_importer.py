import json
import re
import sys
from copy import deepcopy
from datetime import datetime
from pathlib import Path

from docx import Document
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / "docs/source/临商银行十五五战略规划.xlsx"
DOCX = ROOT / "docs/source/临商银行十五五战略规划任务卡片.docx"
PUBLIC_DATA = ROOT / "public/data"
SHARED_DATA = ROOT / "shared/generated-data"
MINI_DATA = ROOT / "miniprogram/data"

TASK_STATUSES = ["待启动", "进行中", "待验收", "已完成", "已逾期", "已暂停"]
RISKS = ["无风险", "低风险", "中风险", "高风险"]
REVIEW_STATUSES = ["草稿", "待审核", "已通过", "已退回"]
STANDARD_STATUSES = ["未开始", "进行中", "待确认", "已达标", "未达标"]
ROLES = [
    {"account": "strategy", "roleCode": "strategy", "roleName": "战略部门", "name": "战略部门", "role": "战略部门", "department": "战略管理部", "description": "全行战略驾驶舱与统筹管理视角", "departments": ["战略管理部"], "permissions": ["all", "review", "export"], "password": "Linshang@2026"},
    {"account": "corporate", "roleCode": "corporate", "roleName": "公司业务部门", "name": "公司业务部门", "role": "公司业务部门", "department": "公司业务部", "description": "对公战略任务牵头部门", "departments": ["公司业务部", "机构业务部", "国际业务部", "票据业务中心", "供应链金融中心", "公司业务部-票据业务中心"], "permissions": ["department"], "password": "Linshang@2026"},
    {"account": "retail", "roleCode": "retail", "roleName": "零售金融部门", "name": "零售金融部门", "role": "零售金融部门", "department": "个人业务部", "description": "零售与财富管理任务牵头部门", "departments": ["个人业务部", "消费金融部", "零售金融部"], "permissions": ["department"], "password": "Linshang@2026"},
    {"account": "inclusive", "roleCode": "inclusive", "roleName": "普惠金融部门", "name": "普惠金融部门", "role": "普惠金融部门", "department": "普惠金融部", "description": "普惠客群和小微业务任务牵头部门", "departments": ["普惠金融部", "小微业务部"], "permissions": ["department"], "password": "Linshang@2026"},
    {"account": "market", "roleCode": "market", "roleName": "金融市场部门", "name": "金融市场部门", "role": "金融市场部门", "department": "金融市场部", "description": "金融市场与同业协同任务部门", "departments": ["金融市场部", "票据业务中心"], "permissions": ["department"], "password": "Linshang@2026"},
    {"account": "risk", "roleCode": "risk", "roleName": "风险合规部门", "name": "风险合规部门", "role": "风险合规部门", "department": "风险管理部", "description": "风险、合规、内控相关任务部门", "departments": ["风险管理部", "法律合规部", "授信审批部", "资产保全部", "内控合规部"], "permissions": ["department"], "password": "Linshang@2026"},
    {"account": "digital", "roleCode": "digital", "roleName": "数字科技部门", "name": "数字科技部门", "role": "数字科技部门", "department": "数字金融部", "description": "数字化和科技建设任务部门", "departments": ["数字金融部", "金融科技部", "科技部"], "permissions": ["department"], "password": "Linshang@2026"},
    {"account": "finance", "roleCode": "finance", "roleName": "计划财务部门", "name": "计划财务部门", "role": "计划财务部门", "department": "计划财务部", "description": "计划、预算、资源配置相关任务部门", "departments": ["计划财务部", "资产负债管理部"], "permissions": ["department"], "password": "Linshang@2026"},
    {"account": "hr", "roleCode": "hr", "roleName": "人力资源部门", "name": "人力资源部门", "role": "人力资源部门", "department": "人力资源部", "description": "组织、人才、绩效相关任务部门", "departments": ["人力资源部"], "permissions": ["department"], "password": "Linshang@2026"},
    {"account": "governance", "roleCode": "governance", "roleName": "公司治理与综合管理部门", "name": "公司治理与综合管理部门", "role": "公司治理与综合管理部门", "department": "综合管理部", "description": "治理、综合协调和行政保障部门", "departments": ["综合管理部", "董事会办公室", "办公室", "运营管理部", "安全保卫部"], "permissions": ["department"], "password": "Linshang@2026"},
    {"account": "audit", "roleCode": "audit", "roleName": "监督审计部门", "name": "监督审计部门", "role": "监督审计部门", "department": "审计部", "description": "监督、审计和整改跟踪部门", "departments": ["审计部", "纪检监察部", "监事会办公室"], "permissions": ["department"], "password": "Linshang@2026"},
    {"account": "branch", "roleCode": "branch", "roleName": "分支机构", "name": "分支机构", "role": "分支机构", "department": "分支机构", "description": "分支机构执行与反馈视角", "departments": ["分支机构", "各分支机构", "各支行"], "permissions": ["department"], "password": "Linshang@2026"},
]


def cell(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    return str(value).strip()


def month(value):
    if value is None or value == "":
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m")
    raw = str(value).strip().replace("年", ".").replace("月", "").replace("-", ".")
    match = re.search(r"(20\d{2})[.](\d{1,2})", raw)
    if match:
        return f"{match.group(1)}-{match.group(2).zfill(2)}"
    if re.fullmatch(r"20\d{2}", raw):
        return f"{raw}-12"
    if re.fullmatch(r"20\d{2}\.\d+", raw):
        year, mon = raw.split(".", 1)
        return f"{year}-{mon.zfill(2)}"
    return raw


def split_departments(text):
    raw = cell(text)
    if not raw:
        return []
    parts = re.split(r"[、,，/；;]\s*", raw)
    return [part.strip() for part in parts if part.strip()]


PREFIX = re.compile(r"^\s*(?:\d+[\.\、．)]|[（(]\d+[）)]|[一二三四五六七八九十]+[、.．])\s*")


def first_sentence(text):
    clean = cell(text)
    match = re.search(r"[。；;]", clean)
    if match:
        clean = clean[: match.end()]
    return clean[:42] + ("..." if len(clean) > 42 else "")


def split_numbered_text(text):
    source = cell(text).replace("\r", "").strip()
    if not source:
        return []
    items = []
    for line in [i.strip() for i in source.split("\n") if i.strip()]:
        if PREFIX.match(line):
            clean = PREFIX.sub("", line).strip()
            items.append({"title": first_sentence(clean), "description": clean, "rawText": line})
        elif items:
            items[-1]["description"] = (items[-1]["description"] + "\n" + line).strip()
            items[-1]["rawText"] = (items[-1]["rawText"] + "\n" + line).strip()
        else:
            items.append({"title": first_sentence(line), "description": line, "rawText": line})
    return items or [{"title": first_sentence(source), "description": source, "rawText": source}]


def split_standards_text(text):
    source = cell(text).replace("\r", "").strip()
    if not source:
        return []
    items = []
    group_name = ""
    for line in [i.strip() for i in source.split("\n") if i.strip()]:
        if not PREFIX.match(line) and re.search(r"[:：]$", line):
            group_name = re.sub(r"[:：]$", "", line).strip()
            continue
        if PREFIX.match(line):
            clean = PREFIX.sub("", line).strip()
            items.append({"groupName": group_name, "title": first_sentence(clean), "description": clean, "rawText": line})
        elif items:
            items[-1]["description"] = (items[-1]["description"] + "\n" + line).strip()
            items[-1]["rawText"] = (items[-1]["rawText"] + "\n" + line).strip()
        else:
            items.append({"groupName": group_name, "title": first_sentence(line), "description": line, "rawText": line})
    return items or [{"groupName": group_name, "title": first_sentence(source), "description": source, "rawText": source}]


def infer_type(text):
    value = cell(text)
    if re.search(r"[0-9]+(?:\.[0-9]+)?\s*(亿元|万元|户|%|％|条|个|项|人|万户)", value):
        return "定量指标"
    if re.search(r"制度|机制|办法|规则|方案|流程|政策", value):
        return "制度或方案"
    if re.search(r"上线|系统|平台", value):
        return "系统上线"
    if re.search(r"产品|服务体系|产品体系", value):
        return "产品上线"
    if re.search(r"团队|人员|岗位|人才|队伍", value):
        return "队伍建设"
    if re.search(r"覆盖|全覆盖|落地|推广", value):
        return "业务覆盖"
    if re.search(r"20\d{2}[.年-]\d{1,2}|阶段|年度", value):
        return "里程碑"
    return "其他"


def target_meta(text):
    value = cell(text)
    date_match = re.search(r"20\d{2}[.年-]\d{1,2}", value)
    number_match = re.search(r"([0-9]+(?:\.[0-9]+)?)\s*(亿元|万元|万户|户|%|％|条|个|项|人)?", value)
    return {
        "targetDate": month(date_match.group(0)) if date_match else "",
        "targetValue": float(number_match.group(1)) if number_match else "",
        "unit": number_match.group(2) if number_match and number_match.group(2) else "",
    }


def parse_word_cards():
    if not DOCX.exists():
        return {}
    cards = {}
    doc = Document(DOCX)
    for table in doc.tables:
        kv = {}
        for row in table.rows:
            if len(row.cells) >= 2:
                key = cell(row.cells[0].text)
                value = cell(row.cells[1].text)
                if key:
                    kv[key] = value
        code = kv.get("编号")
        if code:
            cards[code] = kv
    return cards


def task_progress_seed(index):
    return [18, 35, 48, 62, 76, 90][index % 6]


def risk_seed(index):
    return RISKS[[1, 1, 2, 0, 3, 1, 2][index % 7]]


def review_seed(index):
    return REVIEW_STATUSES[[0, 1, 2, 3, 2, 1][index % 6]]


def build_tasks(wb, word_cards):
    ws = wb["任务卡片反馈版"]
    tasks = []
    seen = {}
    duplicates = []
    empty_required = []
    invalid_dates = []

    for row_no, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
        if not cell(row[1]) and not cell(row[3]):
            continue
        code = cell(row[1]) or f"ROW-{row_no}"
        if code in seen:
            duplicates.append({"code": code, "rows": [seen[code], row_no]})
            code = f"{code}__row{row_no}"
        seen[code] = row_no

        original_code = cell(row[1]) or code
        word = word_cards.get(original_code, {})
        measures_text = cell(row[11]) or word.get("关键实施举措", "")
        excel_standards_text = cell(row[12])
        word_standards_text = word.get("完成标准", "") or word.get("任务完成标准", "")
        excel_standards = split_standards_text(excel_standards_text)
        word_standards = split_standards_text(word_standards_text)
        standards_text = word_standards_text if len(word_standards) > len(excel_standards) else excel_standards_text or word_standards_text
        lead = cell(row[5]) or word.get("牵头部门", "")
        collaborators = split_departments(cell(row[6]) or word.get("协同部门", ""))
        start = month(row[9]) or month(word.get("启动时间", ""))
        end = month(row[10]) or month(word.get("结束时间", ""))
        if start and not re.match(r"20\d{2}-\d{2}$", start):
            invalid_dates.append({"code": code, "field": "startDate", "value": start})
        if end and not re.match(r"20\d{2}-\d{2}$", end):
            invalid_dates.append({"code": code, "field": "endDate", "value": end})

        missing = [name for name, value in {
            "code": original_code,
            "taskName": cell(row[3]) or word.get("任务名称", ""),
            "leadDepartment": lead,
            "originalMeasuresText": measures_text,
            "originalStandardsText": standards_text,
        }.items() if not value]
        if missing:
            empty_required.append({"code": code, "row": row_no, "fields": missing})

        parsed_measures = split_numbered_text(measures_text)
        parsed_standards = split_standards_text(standards_text)
        base_progress = task_progress_seed(len(tasks))
        weight = round(100 / len(parsed_measures), 2) if parsed_measures else 100
        measures = []
        for idx, item in enumerate(parsed_measures, start=1):
            measure_progress = min(100, max(0, base_progress + (idx - 1) * 7 - 10))
            status = "已完成" if measure_progress >= 100 else ("进行中" if measure_progress > 0 else "待启动")
            measures.append({
                "id": f"{code}-M{idx:02d}",
                "taskId": code,
                "order": idx,
                "groupName": "",
                "title": item["title"],
                "description": item["description"],
                "rawText": item["rawText"],
                "ownerDepartment": lead,
                "collaboratorDepartments": collaborators,
                "responsiblePerson": "演示填报人",
                "startDate": start,
                "dueDate": end,
                "weight": weight,
                "status": status,
                "reviewStatus": review_seed(idx + len(tasks)),
                "progress": measure_progress,
                "currentPeriodSummary": f"已围绕“{item['title']}”开展阶段推进。",
                "completedWork": "完成资料梳理、部门沟通和阶段性推进。",
                "keyDeliverables": "形成阶段成果清单和问题台账。",
                "nextPlan": "继续推进关键节点，补充佐证材料。",
                "issueDescription": "" if idx % 3 else "跨部门数据口径仍需统一。",
                "riskLevel": risk_seed(idx + len(tasks)),
                "riskDescription": "" if idx % 3 else "进度受协同部门反馈节奏影响。",
                "responseMeasure": "建立周跟踪机制，必要时提请协调。",
                "coordinationRequest": "" if idx % 2 else "请协同部门确认阶段材料。",
                "decisionRequest": "",
                "evidenceFiles": [],
                "updatedBy": "系统初始化",
                "updatedAt": "2026-07-15 09:00",
                "updateHistory": [],
            })

        standards = []
        for idx, item in enumerate(parsed_standards, start=1):
            meta = target_meta(item["description"])
            status = STANDARD_STATUSES[(idx + len(tasks)) % len(STANDARD_STATUSES)]
            standards.append({
                "id": f"{code}-S{idx:02d}",
                "taskId": code,
                "groupName": item.get("groupName", ""),
                "order": idx,
                "description": item["description"],
                "rawText": item["rawText"],
                "type": infer_type(item["description"]),
                "targetValue": meta["targetValue"],
                "unit": meta["unit"],
                "targetDate": meta["targetDate"] or end,
                "currentValue": meta["targetValue"] if status == "已达标" and meta["targetValue"] != "" else "",
                "completionRate": 100 if status == "已达标" else min(95, base_progress + idx * 4),
                "status": status,
                "mandatory": True,
                "linkedMeasureIds": [measures[min(idx - 1, len(measures) - 1)]["id"]] if measures else [],
                "evidenceFiles": [],
                "verificationNote": "Demo 初始化验收项，支持本地登记说明和佐证材料。",
                "reviewer": "演示审核人" if status in ["已达标", "未达标"] else "",
                "reviewedAt": "2026-07-15 09:00" if status in ["已达标", "未达标"] else "",
            })

        task = {
            "id": code,
            "code": original_code,
            "businessSection": cell(row[2]) or word.get("业务板块", ""),
            "chapter": cell(row[2]).split("-")[0] if cell(row[2]) else "",
            "taskName": cell(row[3]) or word.get("任务名称", ""),
            "taskGoal": cell(row[4]) or word.get("任务目标", ""),
            "leadDepartment": lead,
            "collaboratorDepartments": collaborators,
            "priority": cell(row[7]) or "中",
            "startDate": start,
            "endDate": end,
            "implementationPeriod": cell(row[8]) or f"{start} 至 {end}",
            "status": TASK_STATUSES[[0, 1, 1, 2, 3, 4][len(tasks) % 6]],
            "riskLevel": risk_seed(len(tasks)),
            "progress": base_progress,
            "originalMeasuresText": measures_text,
            "originalStandardsText": standards_text,
            "measures": measures,
            "completionStandards": standards,
            "milestones": [
                {"id": f"{code}-ML01", "name": "启动与方案明确", "date": start, "status": "已完成" if base_progress > 35 else "进行中"},
                {"id": f"{code}-ML02", "name": "阶段成果提交", "date": end, "status": "未开始" if base_progress < 70 else "进行中"},
            ],
            "currentPeriodSummary": "本期围绕任务目标完成阶段性推进，重点事项按计划跟踪。",
            "completedWork": "已完成任务拆解、责任确认和阶段资料整理。",
            "keyDeliverables": "任务台账、举措清单、验收项清单。",
            "nextPlan": "继续推进重点举措，完善佐证材料和审核说明。",
            "issueDescription": cell(row[15]),
            "riskDescription": "存在跨部门协同、数据口径或系统建设节奏风险。" if len(tasks) % 4 == 0 else "",
            "responseMeasure": "通过周例会、专项协调和台账督办推进。",
            "coordinationRequest": "需相关部门按期反馈进度。" if len(tasks) % 5 == 0 else "",
            "decisionRequest": "需管理层确认资源配置。" if cell(row[13]) else "",
            "evidenceFiles": [],
            "reviewStatus": review_seed(len(tasks)),
            "createdAt": "2026-07-15 09:00",
            "updatedAt": "2026-07-15 09:00",
            "updatedBy": "系统初始化",
            "updateHistory": [
                {"time": "2026-07-15 09:00", "operator": "系统初始化", "action": "导入任务", "note": "由 Excel/Word 源文件生成静态演示数据。"}
            ],
            "source": {"workbook": XLSX.name, "sheet": ws.title, "row": row_no, "wordMatched": original_code in word_cards},
            "rawText": {"excelRow": [cell(v) for v in row], "wordCard": word},
        }
        tasks.append(task)

    return tasks, {
        "duplicateTaskIds": duplicates,
        "emptyRequiredFields": empty_required,
        "invalidDates": invalid_dates,
    }


def build_indicators(wb):
    ws = wb["考核指标（发送版） 0520"]
    indicators = []
    current_line = ""
    current_department = ""
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        line = cell(row[0]) or current_line
        name = cell(row[1])
        dept = cell(row[2]) or current_department
        if not name:
            continue
        current_line, current_department = line, dept
        values = {str(year): row[pos] for year, pos in [(2025, 3), (2026, 4), (2027, 5), (2028, 6), (2029, 7), (2030, 8)]}
        target = values.get("2026")
        current = target * 0.74 if isinstance(target, (int, float)) else ""
        rate = round(current / target * 100) if isinstance(target, (int, float)) and target else ""
        indicators.append({
            "id": f"KPI-{idx:03d}",
            "businessLine": line,
            "name": name,
            "department": dept,
            "unit": infer_unit(name),
            "values": values,
            "target2026": target,
            "currentValue": round(current, 2) if isinstance(current, (int, float)) else "",
            "achievementRate": rate,
            "status": "达标" if isinstance(rate, int) and rate >= 100 else ("关注" if isinstance(rate, int) and rate >= 80 else "未达标"),
            "riskLevel": "高风险" if isinstance(rate, int) and rate < 70 else ("中风险" if isinstance(rate, int) and rate < 85 else "低风险"),
            "updatedAt": "2026-07-15",
            "source": {"workbook": XLSX.name, "sheet": ws.title, "row": idx},
        })
    return indicators


def infer_unit(name):
    match = re.search(r"（([^）]+)）", name)
    if match:
        return match.group(1)
    if "收入" in name or "规模" in name or "余额" in name:
        return "亿元"
    if "客户" in name:
        return "户"
    return ""


def build_systems(wb):
    ws = wb["系统建设清单"]
    systems = []
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not cell(row[2]):
            continue
        systems.append({
            "id": f"SYS-{len(systems)+1:03d}",
            "name": cell(row[2]),
            "leadDepartment": cell(row[3]),
            "targetDate": month(row[4]),
            "status": TASK_STATUSES[len(systems) % 4],
            "progress": [20, 45, 60, 80][len(systems) % 4],
            "source": {"workbook": XLSX.name, "sheet": ws.title, "row": idx},
        })
    return systems


def build_departments(tasks):
    names = sorted(set([t["leadDepartment"] for t in tasks if t["leadDepartment"]] + [d for t in tasks for d in t["collaboratorDepartments"]]))
    departments = []
    for name in names:
        lead_tasks = [t for t in tasks if t["leadDepartment"] == name]
        collab_tasks = [t for t in tasks if name in t["collaboratorDepartments"]]
        all_tasks = lead_tasks + [t for t in collab_tasks if t not in lead_tasks]
        avg = round(sum(t["progress"] for t in all_tasks) / len(all_tasks)) if all_tasks else 0
        departments.append({
            "id": f"D-{len(departments)+1:03d}",
            "name": name,
            "leadTaskCount": len(lead_tasks),
            "collaborateTaskCount": len(collab_tasks),
            "averageProgress": avg,
            "highRiskCount": sum(1 for t in all_tasks if t["riskLevel"] == "高风险"),
            "overdueCount": sum(1 for t in all_tasks if t["status"] == "已逾期"),
            "submittedCount": sum(1 for t in all_tasks if t["reviewStatus"] in ["待审核", "已通过"]),
            "returnedCount": sum(1 for t in all_tasks if t["reviewStatus"] == "已退回"),
            "fillRate": round(sum(1 for t in all_tasks if t["reviewStatus"] in ["待审核", "已通过"]) / len(all_tasks) * 100) if all_tasks else 0,
        })
    return departments


def write_json(path, data):
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def write_js(path, name, data):
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(f"const {name} = {json.dumps(data, ensure_ascii=False, indent=2)};\nmodule.exports = {name};\n", encoding="utf-8")


def main():
    wb = load_workbook(XLSX, data_only=True)
    word_cards = parse_word_cards()
    tasks, task_report = build_tasks(wb, word_cards)
    indicators = build_indicators(wb)
    systems = build_systems(wb)
    departments = build_departments(tasks)

    report = {
        "generatedAt": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "importedTaskCount": len(tasks),
        "importedIndicatorCount": len(indicators),
        "importedSystemCount": len(systems),
        "wordCardCount": len(word_cards),
        "duplicateTaskIds": task_report["duplicateTaskIds"],
        "emptyRequiredFields": task_report["emptyRequiredFields"],
        "invalidDates": task_report["invalidDates"],
        "parsingWarnings": [
            {"code": t["code"], "message": "关键实施举措未识别到编号，已整体保留为一项。"}
            for t in tasks if len(t["measures"]) == 1 and not PREFIX.match(t["originalMeasuresText"])
        ],
        "sourceConflicts": [
            {"code": t["code"], "message": "Word 与 Excel 均存在任务卡，已优先采用 Excel 反馈版字段。"}
            for t in tasks if t["source"]["wordMatched"]
        ],
        "skippedRows": [],
        "measureCount": sum(len(t["measures"]) for t in tasks),
        "completionStandardCount": sum(len(t["completionStandards"]) for t in tasks),
    }

    for folder in [PUBLIC_DATA, SHARED_DATA]:
        write_json(folder / "tasks.json", tasks)
        write_json(folder / "indicators.json", indicators)
        write_json(folder / "systems.json", systems)
        write_json(folder / "departments.json", departments)
        write_json(folder / "demo-users.json", ROLES)
        write_json(folder / "import-report.json", report)

    write_js(MINI_DATA / "tasks.js", "tasks", tasks)
    write_js(MINI_DATA / "indicators.js", "indicators", indicators)
    write_js(MINI_DATA / "systems.js", "systems", systems)
    write_js(MINI_DATA / "departments.js", "departments", departments)
    write_js(MINI_DATA / "demo-users.js", "demoUsers", ROLES)
    write_js(MINI_DATA / "import-report.js", "importReport", report)

    print(json.dumps(report, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        print(f"Import failed: {exc}", file=sys.stderr)
        raise
